import os
import re
import math
import pandas as pd
import xml.etree.ElementTree as ET
from saxonche import PySaxonProcessor
from collections import defaultdict
from itertools import chain
from pathlib import Path
from tqdm import tqdm
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.comments import Comment

# Function to import xml files and their config files
def import_files(xml_path, config_path, xml_recursive=False, config_recursive=False):
    """
    Imports and processes authority files and their configuration.
    Args:
        xml_path (str): Directory containing XML files.
        config_path (str): Directory containing authority configuration CSV files.
    Returns:
        tuple: A dictionary of parsed XML files, a dictionary of configuration DataFrames, and a dictionary of empty DataFrames.
    """
    # Read and parse XML files
    xml_files = read_files(xml_path, pattern=".xml", recursive=xml_recursive)
    xml_data = {}
    for file in tqdm(xml_files, desc="Parsing XML files"):
        try:
            filename = os.path.splitext(os.path.basename(file))[0]
            xml_data[filename] = parse_xml(file)
        except Exception as e:
            tqdm.write(f"Failed to parse file {file}. Error: {e}")

    # Read and parse CSV configuration files
    config_files = sorted(read_files(config_path, pattern=".csv", recursive=config_recursive))
    config_list = {}
    for file in tqdm(config_files, desc="Parsing config files"):
        try:
            name = os.path.splitext(os.path.basename(file))[0]
            config_list[name] = pd.read_csv(file, dtype='str', na_values=["", "nan", "NaN"]).where(pd.notna, None)
        except Exception as e:
            tqdm.write(f"Failed to parse config file {file}. Error: {e}")

    # Create an empty DataFrame for each configuration file
    df_list = {}

    for name, config in config_list.items():
        columns = []
        seen = defaultdict(int)
        
        for col in zip(config['section'], config['heading']):
            base_name = col[0] + ": " + col[1] if col[0] is not None else col[1]
            if seen[base_name] > 0:
                new_name = f"{base_name}_{seen[base_name]}"
                tqdm.write(f"Warning: Duplicate column name '{base_name}' detected. Renaming to '{new_name}'.")
            else:
                new_name = base_name
            columns.append(new_name)
            seen[base_name] += 1
        
        df_list[name] = pd.DataFrame(columns=columns)

    return xml_data, config_list, df_list

# Helper function to read files from a directory
def read_files(directory, pattern, recursive=True):
    """
    Reads files from a specified directory.
    Args:
        directory (str): Directory to search for files.
        pattern (str): File extension pattern to match.
        recursive (bool): Whether to search subdirectories.
    Returns:
        list: List of file paths.
    """
    try:
        directory_path = Path(directory)
        if recursive:
            files = list(directory_path.rglob(f"*{pattern}"))
        else:
            files = list(directory_path.glob(f"*{pattern}"))
        return [str(file) for file in files]
    except Exception as e:
        tqdm.write(f"Reading files in {directory} failed. Error: {e}")
        raise

# Helper function to parse an XML file and return its root element
def parse_xml(file):
    """
    Parses an XML file and returns the root element.
    Args:
        file (str): Path to the XML file.
    Returns:
        Element: The root element of the parsed XML file.
    """
    try:
        tree = ET.parse(file)
        root = tree.getroot()
        return root
    except Exception as e:
        tqdm.write(f"Parsing {file} failed. Error: {e}")
        raise

# Function to extract data from the XML files based on the configuration files
def process_file(
    file_type,
    config_name,
    config,
    xml_data,
    df_list,
    csv_output_dir,
    json_output_dir,
    separator_map=None,
    lookup_df_list=None,
    bar_pos=1,
    cores_spare=0
):
    """
    Function that processes either authority or collection files depending on file_type.
    Args:
        file_type (str): Either 'authority' or 'collection' to choose which branch to run.
        config_name (str): Name of the configuration file.
        config (DataFrame): The configuration DataFrame.
        xml_data (dict): Dictionary of XML data.
        df_list (dict): Dictionary of DataFrames keyed by configuration name.
        csv_output_dir (str): Output directory for CSV files.
        json_output_dir (str): Output directory for JSON files.
        separator_map (dict): Dictionary of separators for authority lookups. Default None.
        lookup_df_list (dict): Dictionary of DataFrames for authority XML files; used for collection branch. Default None.
        bar_pos (int): Position parameter for tqdm progress bar. Default 1.
    Returns:
        tuple: (config_name, processed DataFrame)
    """
    try:
        df = df_list[config_name]
        # Extract config columns
        if file_type == "authority":
            try:
                auth_files, xpaths, formats = (
                    config[col].tolist() for col in ["auth_file", "xpath", "format"]
                )
            except Exception as e:
                tqdm.write(f"Failed to extract configuration columns for '{config_name}'. Error: {e}")
                return config_name, df
            num_workers = max(1, (os.cpu_count() or 1) - cores_spare)
            all_args = [
                (i, xpath, auth_file)
                for i, (xpath, auth_file) in enumerate(zip(xpaths, auth_files))
            ]
        elif file_type == "collection":
            try:
                xpaths, auth_files, auth_sections, auth_cols, separators, formats = (
                    config[col].tolist() for col in ["xpath", "auth_file", "auth_section", "auth_col", "separator", "format"]
                )
            except Exception as e:
                tqdm.write(f"Failed to extract configuration columns for '{config_name}'. Error: {e}")
                return config_name, df
            num_workers = max(1, (os.cpu_count() or 1) - cores_spare)
            all_args = [
                (i, xpath, auth_file, auth_section, auth_col, separator)
                for i, (xpath, auth_file, auth_section, auth_col, separator)
                in enumerate(zip(xpaths, auth_files, auth_sections, auth_cols, separators))
            ]
        else:
            raise ValueError(f"Unsupported file_type: '{file_type}'. Valid options are 'authority' or 'collection'.")

        # Batch processing
        batch_size = max(1, math.ceil(len(all_args) / (num_workers + 1)))
        batches = [all_args[i:i+batch_size] for i in range(0, len(all_args), batch_size)]
        futures = []
        with ProcessPoolExecutor(max_workers=num_workers) as executor:
            for batch in batches:
                if file_type == "authority":
                    futures.append(executor.submit(process_batch, batch=batch, file_type=file_type, xml_data=xml_data))
                else:
                    futures.append(executor.submit(process_batch, batch=batch, file_type=file_type, xml_data=xml_data, lookup_df_list=lookup_df_list, separator_map=separator_map))
            for future in tqdm(as_completed(futures), total=len(futures), desc=f"File '{config_name}'", position=bar_pos):
                try:
                    batch_result = future.result()
                    for i, col_data in batch_result.items():
                        # Normalise all cells to scalars
                        df.iloc[:, i] = [normalise_cell(x) for x in col_data]
                except Exception as e:
                    tqdm.write(f"Error processing batch for '{config_name}'. Error: {e}")
    except Exception as e:
        tqdm.write(f"Failed to fully process '{config_name}'. Returning an empty DataFrame. Error: {e}")
        return config_name, pd.DataFrame()

    # Defragment, unlist, format, and sort
    df = defrag(df)
    df = unlist_columns(df)
    df = set_format(df, formats=formats)
    df = sort_df(df=df, file_type=file_type)
    save_as(df, csv_output_dir, config_name, format="csv")
    save_as(df, json_output_dir, config_name, format="json")
    return config_name, df

# Helper function to normalise any cell to a scalar string
def normalise_cell(cell):
    """ Normalises a cell to a scalar string.
    Args:
        cell: The cell value to normalise.
    Returns:
        str: The normalised cell value as a string.
    """
    if isinstance(cell, list):
        flat = list(chain.from_iterable(cell)) if any(isinstance(i, list) for i in cell) else cell
        cell_text = "".join(str(x) if x is not None else "" for x in flat)
        tqdm.write(f"Warning: Cell values were received as a list and have been merged to: '{cell_text}'.")
        return cell_text
    elif cell is None or (isinstance(cell, float) and math.isnan(cell)):
        return ""
    else:
        return str(cell)

# Helper function to process batches of columns
def process_batch(
    batch,
    file_type,
    xml_data,
    lookup_df_list=None,
    separator_map=None
):
    """
    Processes a batch of columns for either authority or collection files.
    Args:
        batch (list): List of tuples containing column processing arguments.
        file_type (str): Either 'authority' or 'collection'.
        xml_data (dict): Dictionary of XML files.
        lookup_df_list (dict, optional): Dictionary of DataFrames for authority files (for collection).
        separator_map (dict, optional): Dictionary of separators for authority lookups (for collection).
    Returns:
        dict: Dictionary of processed results.
    """
    out = {}
    for args in batch:
        try:
            # Unpack args according to file_type
            if file_type == "authority":
                i, xpath, auth_file = args
                result = process_column(
                    i,
                    xpath,
                    auth_file,
                    xml_data,
                    file_type=file_type,
                    lookup_df_list=lookup_df_list,
                    separator_map=separator_map
                )
            elif file_type == "collection":
                i, xpath, auth_file, auth_section, auth_col, separator = args
                result = process_column(
                    i,
                    xpath,
                    auth_file,
                    xml_data,
                    lookup_df_list=lookup_df_list,
                    auth_section=auth_section,
                    auth_col=auth_col,
                    separator=separator,
                    separator_map=separator_map,
                    file_type=file_type
                )
            else:
                raise ValueError(f"Unsupported file_type: '{file_type}'. Valid options are 'authority' or 'collection'.")

            if result is not None:
                i, col_data = result
                out[i] = col_data
        except Exception as e:
            tqdm.write(f"Failed to process column batch. Error: {e}")
            continue

    return out

# Helper function to process a single column
def process_column(
    i,
    xpath,
    auth_file,
    xml_data,
    lookup_df_list=None,
    auth_section=None,
    auth_col=None,
    separator=None,
    separator_map=None,
    file_type=None
):
    """
    Processes a single column for either authority or collection DataFrame.
    Args:
        i (int): The index of the column.
        xpath (str): The XPath expression to extract data.
        auth_file (str): The authority file name.
        xml_data (dict): Dictionary of XML files.
        lookup_df_list (dict, optional): Dictionary of DataFrames for authority files (for collection).
        auth_section (str, optional): The section name in the authority file (for collection).
        auth_col (str, optional): The column name in the authority file (for collection).
        separator (str, optional): The separator for joining values (for collection).
        separator_map (dict, optional): Dictionary of separators for authority lookups (for collection).
        file_type (str): Either 'authority' or 'collection'.
    Returns:
        tuple: The index and the processed results.
    """
    if file_type == "authority":
        auth_xml = xml_data.get(auth_file)
        results = extract_with_xpath(auth_xml, xpath)

        # Flatten the results and ensure no nested lists remain
        if results is not None:
            results = list(chain.from_iterable(sublist if isinstance(sublist, list) else [sublist] for sublist in results))

            if not auth_file or (lookup_df_list is not None and auth_file.lower().strip() not in lookup_df_list):
                results = [item for item in results if not isinstance(item, list)]
            return i, results

    elif file_type == "collection":
        results = []
        if not auth_file or (lookup_df_list is not None and auth_file.lower().strip() not in lookup_df_list):
            pass
        # If auth_file is not in lookup_df_list keys, extract the data and append directly
        if not auth_file or auth_file.lower().strip() not in (lookup_df_list or {}):
            for filename, xml in xml_data.items():
                results.append(extract_with_xpath(xml, xpath))
        else:
            # Set up auth lookup DataFrame
            auth_df = None
            if lookup_df_list is not None:
                auth_df = lookup_df_list.get(auth_file.lower().strip())

            # Set the separator
            s = get_separator(separator, separator_map)

            # Set the column name, handling None values
            auth_section_str = auth_section + ": " if auth_section is not None else ""
            auth_col_str = auth_col if auth_col is not None else ""
            auth_col_name = auth_section_str + auth_col_str

            # Lookup the value in the authority file
            for filename, xml in xml_data.items():
                # Extract data items using XPath
                extracted_data = extract_with_xpath(xml, xpath)
                
                # Process each data item using the lookup function
                lookup_data = []
                if extracted_data is not None:
                    for data_item in extracted_data:
                        processed_item = process_lookup_item(data_item, auth_df, auth_col_name, s)
                        # Flatten and validate processed_item before appending
                        if isinstance(processed_item, list):
                            lookup_data.extend(processed_item)
                        else:
                            lookup_data.append(processed_item)
                    results.append(lookup_data)

        # Flatten the results and ensure no nested lists remain
        results = [item for sublist in results for item in (sublist if isinstance(sublist, list) else [sublist])]
        return i, results

    else:
        raise ValueError(f"Unsupported file_type: {file_type}. Valid options are 'authority' or 'collection'.")

# Helper function to apply XPath 3.1 queries to an XML element in the TEI namespace
def extract_with_xpath(xml_element, xpath_expr):
    """
    Extracts data from an XML element using XPath 3.1 queries.
    Args:
        xml_element (Element): The XML element to search.
        xpath_expr (str): The XPath expression to evaluate.
    Returns:
        list: The extracted data.
    """
    if xml_element is None:
        tqdm.write("XPath extraction failed. XML file not found.")
        return []

    try:
        xml_str = ET.tostring(xml_element, encoding="unicode")
        with PySaxonProcessor(license=False) as proc:
            doc = proc.parse_xml(xml_text=xml_str)
            if doc is None:
                tqdm.write("Failed to parse XML string into Saxon XdmNode.")
                return []

            xpath_proc = proc.new_xpath_processor()
            xpath_proc.declare_namespace("tei", "http://www.tei-c.org/ns/1.0")
            xpath_proc.set_context(xdm_item=doc)

            result = xpath_proc.evaluate(xpath_expr)
            if result is None:
                return []

            # Convert result to a list of strings
            items = []
            try:
                for i in range(result.size):
                    item = result.item_at(i)
                    if hasattr(item, 'string_value'):
                        items.append(item.string_value)
                    else:
                        items.append(str(item))
            except Exception as e:
                tqdm.write(f"Could not iterate result items. Error: {e}")
                # fallback for singleton result
                try:
                    if hasattr(result, 'string_value'):
                        items.append(result.string_value)
                    else:
                        items.append(str(result))
                except Exception as inner_e:
                    tqdm.write(f"Failed to extract string value from result. Error: {inner_e}")
                    return []

            return items

    except Exception as e:
        tqdm.write(f"XPath extraction failed. Offending XPath: {xpath_expr}. Error: {e}")
        return []

# Helper function to determine the separator for authority lookups
def get_separator(separator, separator_map):
    """
    Determines the separator to use for authority lookups based on the provided separator and a separator map.
    Args:
        separator (str): The separator to use.
        separator_map (dict): A dictionary mapping separators to their string representations.
    Returns:
        str: The determined separator string."""
    if separator_map is None:
        s = "; "
        tqdm.write(f"No separator map found. Using '{s}' instead.")
    elif str(separator).lower().strip() in separator_map:
        s = separator_map.get(str(separator).lower().strip())
    elif "default" in separator_map:
        s = separator_map.get("default")
        tqdm.write(f"Encountered unexpected separator '{separator}'. Using default '{s}' instead.")
    else:
        s = "; "
        tqdm.write(f"Encountered unexpected separator '{separator}' and no default found. Using '{s}' instead.")
    return s

# Helper function to process data found through the authority lookup
def process_lookup_item(data_item, auth_df, auth_col_name, separator):
    """
    Processes a single data item by looking it up in the authority DataFrame and returning the corresponding value.
    Args:
        data_item (str): The data item to process.
        auth_df (DataFrame): The authority DataFrame.
        auth_col_name (str): The column name in the authority DataFrame.
        separator (str): The separator for joining values.
    Returns:
        str: The processed value, joined by the separator.
    """
    pieces = []
    # Split the data_item on spaces
    for identifier in data_item.split(" "):
        if auth_df is not None:
            # Filter the DataFrame rows where the first column equals the identifier
            filtered = auth_df[auth_df.iloc[:, 0] == identifier]
            if not filtered.empty:
                # Get the value from the specified column
                value = filtered[auth_col_name].iloc[0]
                # If the value is a boolean, convert its string form to lowercase
                piece = str(value).lower() if isinstance(value, bool) else str(value)
            else:
                piece = ""
        else:
            piece = ""
        # Add only non-empty strings
        if piece:
            pieces.append(piece)
    
    # Deduplicate preserving the order by leveraging dict.fromkeys
    deduped = list(dict.fromkeys(pieces))
    # If all strings were empty, deduped will be empty; return a single empty string
    return separator.join(deduped) if deduped else ""

# Helper function to defragment the DataFrame
def defrag(df):
    """
    Defragment the DataFrame columns.
    Args:
        df (DataFrame): The DataFrame to defragment.
    Returns:
        DataFrame: The defragmented DataFrame.
    """
    new_df = df.copy()
    return new_df

# Helper function to unlist columnd in DataFrame
def unlist_columns(df):
    """
    Unlists each cell in the DataFrame (ensures only one value per cell, or joins lists as string).
    Args:
        df (DataFrame): The DataFrame to unlist.
    Returns:
        DataFrame: The unlisted DataFrame.
    """
    for col in df.columns:
        try:
            df[col] = df[col].apply(normalise_cell)
        except Exception as e:
            tqdm.write(f"Failed to unlist column '{col}'. Error: {e}")
    return df

# Helper function to set data formats in DatFrame
def set_format(df, formats):
    """
    Set the format of the DataFrame columns based on the provided formats.
    Args:
        df (DataFrame): The DataFrame to format.
        formats (Series): Series containing the format for each column.
    Returns:
        DataFrame: The formatted DataFrame.
    """
    try:
        for i, col in enumerate(df.columns):
            if formats[i] == "text":
                df[col] = df[col].astype(str)
            elif formats[i] == "number":
                numeric_series = pd.to_numeric(df[col], errors='coerce')
                df[col] = df[col].where(numeric_series.isna(), numeric_series)
            elif formats[i] == "date":
                new_col = []
                for val in df[col]:
                    if pd.isna(val):
                        new_col.append(val)
                    else:
                        val_str = str(val).strip()
                        if val_str.isdigit() or (val_str.startswith('-') and val_str[1:].isdigit()):
                            new_col.append(int(val_str))
                        else:
                            try:
                                dt = pd.to_datetime(val, errors='raise', yearfirst=True)
                                if dt.year < 1900:
                                    new_col.append(val)
                                else:
                                    new_col.append(dt.date())
                            except (ValueError, TypeError):
                                new_col.append(val)
                df[col] = pd.Series(new_col, index=df.index)
            elif formats[i] == "boolean":
                new_col = []
                for val in df[col]:
                    if pd.isna(val):
                        new_col.append(val)
                    else:
                        val_str = str(val).strip().lower()
                        if val_str in ['true', '1', 'yes']:
                            new_col.append(True)
                        elif val_str in ['false', '0', 'no']:
                            new_col.append(False)
                        else:
                            new_col.append(val)
                df[col] = pd.Series(new_col, index=df.index)
            elif formats[i] == "percentage":
                new_col = []
                for val in df[col]:
                    try:
                        num = float(val)
                        if not pd.isna(num):
                            num = min(num, 1.0)
                            new_col.append(f"{num * 100:.2f}%")
                        else:
                            new_col.append(val)
                    except (ValueError, TypeError):
                        new_col.append(val)
                df[col] = pd.Series(new_col, index=df.index)
            else:
                df[col] = df[col].astype(str)
                tqdm.write(f"Unknown format '{formats[i]}'. Column '{col}' will be formatted as text.")
    except Exception as e:
        tqdm.write(f"Error occurred while setting formats. Error: {e}")
    return df

# Helper function to sort data frames
def sort_df(df, file_type):
    """
    Sorts the DataFrame based on file_type ('authority' or 'collection').
    Args:
        df (DataFrame): The DataFrame to sort.
        file_type (str): Either 'authority' or 'collection'.
    Returns:
        DataFrame: The sorted DataFrame.
    """
    try:
        if file_type == "authority":
            if df.iloc[:, 0].astype(str).str.contains(r'_\d+', na=False).any():
                df['temp'] = df.iloc[:, 0].astype(str).str.extract(r'_(\d+)', expand=False).astype(float)
                df = df.sort_values(by='temp', ascending=True, na_position='last').reset_index(drop=True)
                df.drop(columns='temp', inplace=True)
            else:
                df = df.sort_values(by=df.columns[0], ascending=True, na_position='last').reset_index(drop=True)

        elif file_type == "collection":
            if 'metadata: collection' in df.columns and 'metadata: shelfmark' in df.columns:
                df = df.assign(
                    _collection_sort=df['metadata: collection'],
                    _shelfmark_sort=df['metadata: shelfmark'].map(parse_shelfmark)
                ).sort_values(
                    by=['_collection_sort', '_shelfmark_sort'],
                    ascending=True,
                    na_position='last'
                ).drop(columns=['_collection_sort', '_shelfmark_sort']).reset_index(drop=True)

            elif 'metadata: collection' in df.columns:
                first_col = df.columns[0]
                df = df.assign(
                    _collection_sort=df['metadata: collection'],
                    _first_sort=df[first_col].map(parse_shelfmark)
                ).sort_values(
                    by=['_collection_sort', '_first_sort'],
                    ascending=True,
                    na_position='last'
                ).drop(columns=['_collection_sort', '_first_sort']).reset_index(drop=True)

            elif 'metadata: shelfmark' in df.columns:
                first_col = df.columns[0]
                df = df.assign(
                    _shelfmark_sort=df['metadata: shelfmark'].map(parse_shelfmark),
                    _first_sort=df[first_col]
                ).sort_values(
                    by=['_shelfmark_sort', '_first_sort'],
                    ascending=True,
                    na_position='last'
                ).drop(columns=['_shelfmark_sort', '_first_sort']).reset_index(drop=True)

            else:
                first_col = df.columns[0]
                df = df.assign(
                    _first_sort=df[first_col]
                ).sort_values(
                    by=['_first_sort'],
                    ascending=True,
                    na_position='last'
                ).drop(columns=['_first_sort']).reset_index(drop=True)

        else:
            tqdm.write(f"Unknown file_type '{file_type}' for sorting. Returning unsorted DataFrame.")

        return df

    except Exception as e:
        tqdm.write(f"Sorting DataFrame failed. Error: {e}. Returning unsorted DataFrame.")

        return df

# Helper function for natural sorting of shelfmarks
def parse_shelfmark(text):
    """
    Converts a shelfmark into a tuple for deterministic natural sorting.
    Args:
        text (str): The shelfmark string to parse.
    Returns:
        tuple: Components for sorting (ints for numbers, strings for text).
    """
    if pd.isnull(text):
        return ()

    # Normalise dashes
    clean = str(text).replace('–', '-').replace('—', '-')

    # Split into tokens on non-word boundaries
    tokens = re.split(r'[^\w]+', clean)
    parsed = []

    for token in tokens:
        if not token:
            continue

        # Handle dash ranges like "10-20"
        if re.match(r'^\d+-\d+$', token):
            start, end = map(int, token.split('-'))
            sort_value = end if start <= end else start
            parsed.append(sort_value)

        # Handle digit+letter suffix like "10b"
        elif re.match(r'^\d+[a-zA-Z]$', token):
            parsed.append(int(token[:-1]))
            parsed.append(token[-1].lower())

        # Handle digit+letter dash ranges like "10a-b"
        elif re.match(r'^\d+[a-zA-Z]+-[a-zA-Z]+$', token):
            parsed.append(token.split('-'))

        # Handle plain digits
        elif token.isdigit():
            parsed.append(int(token))

        else:
            parsed.append(token.lower())

        parsed.append("")

    return tuple(parsed)

# Helper function to save DataFrame as either csv or json file
def save_as(df, output_dir, config_name, format):
    """
    Saves a DataFrame to a file in the specified format.
    Args:
        df (DataFrame): The DataFrame to save.
        output_dir (str): Directory to save the file.
        config_name (str): Name of the configuration file.
        format (str): File format to save as. Must be 'csv' or 'json'.
    """
    os.makedirs(output_dir, exist_ok=True)
    output_filename = f"{config_name}.{format}"
    output_file = os.path.join(output_dir, output_filename)

    if df.empty:
        tqdm.write(f"Skipping saving '{config_name}' as the DataFrame is empty.")
        return

    try:
        if format == "csv":
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
            tqdm.write(f"Saved '{config_name}' to '{output_file}'")
        elif format == "json":
            df.to_json(output_file, orient='records', force_ascii=False)
            tqdm.write(f"Saved '{config_name}' to '{output_file}'")
        else:
            tqdm.write(f"Invalid format '{format}'. Supported formats are 'csv' and 'json'.")
    except Exception as e:
        tqdm.write(f"Saving data to '{output_filename}' failed. Error: {e}")

# Helper function to save DataFrame list as an xlsx file with individual tables as tabs
def save_as_xlsx(df_list, config_list, output_dir, output_filename):
    """
    Saves a list of DataFrames to an Excel file with each DataFrame in a separate sheet.
    Args:
        df_list (dict): Dictionary of DataFrames to save.
        config_list (dict): Dictionary of configuration DataFrames for headings and sections.
        output_dir (str): Directory to save the Excel file.
        output_filename (str): Name of the output Excel file (without extension).
    """
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"{output_filename}.xlsx")
    sections_list = [config_list[config_name]['section'].to_numpy() for config_name in config_list.keys()]
    headings_list = [config_list[config_name]['heading'].to_numpy() for config_name in config_list.keys()]
    comments_list = [config_list[config_name]['comment'].to_numpy() for config_name in config_list.keys()]
    tqdm.write(f"Saving '{output_filename}.xlsx'...")

    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for (name, df), sections, headings, comments in zip(df_list.items(), sections_list, headings_list, comments_list):
                if df.empty:
                    tqdm.write(f"Skipping saving '{name}' as the DataFrame is empty.")
                    return
                
                # Reformat booleans
                for col in df.select_dtypes(include="bool"):
                    df[col] = df[col].astype(bool)

                # Write DataFrame starting from row 3 (leave room for section + heading rows)
                df.to_excel(writer, sheet_name=name, index=False, startrow=2, header=False)
                worksheet = writer.sheets[name]

                # If values end with '%', format as percentage
                percentage_pattern = re.compile(r'^\d+\.?\d*%$')
                for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        if isinstance(cell.value, str) and percentage_pattern.match(cell.value):
                            try:
                                cell.value = float(cell.value.rstrip('%')) / 100
                                cell.number_format = '0.00%'
                            except ValueError:
                                pass

                # Define styling
                fill = PatternFill("solid", fgColor="082144")
                bold_white_font = Font(bold=True, color="FFFFFF")
                center_align = Alignment(horizontal="center", vertical="center")
                thin_side = Side(style="thin")

                # === Row 1: Section Titles ===
                for col_idx, value in enumerate(sections, start=1):
                    cell = worksheet.cell(row=1, column=col_idx, value=value)
                    cell.fill = fill
                    cell.font = bold_white_font
                    cell.alignment = center_align

                # Merge and center identical adjacent section titles
                start = 0
                for i in range(1, len(sections) + 1):
                    if i == len(sections) or sections[i] != sections[start]:
                        if i - start > 1:
                            worksheet.merge_cells(
                                start_row=1, start_column=start + 1,
                                end_row=1, end_column=i
                            )
                        # Apply style to all cells in merged block
                        for j in range(start + 1, i + 1):
                            cell = worksheet.cell(row=1, column=j)
                            cell.fill = fill
                            cell.font = bold_white_font
                            cell.alignment = center_align
                        start = i

                # === Row 2: Column Headings ===
                for col_idx, value in enumerate(headings, start=1):
                    cell = worksheet.cell(row=2, column=col_idx, value=value)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Add comments to headings
                for col_idx, comment_text in enumerate(comments, start=1):
                    cell = worksheet.cell(row=2, column=col_idx)
                    comment = Comment(comment_text, "Generated")
                    num_lines = (len(str(comment_text)) // 15)
                    comment.height = 30 + 15 * num_lines
                    comment.width = 200
                    cell.comment = comment

                # Force formula-looking strings to text in data cells
                for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith(('=', '-', '+', '@')):
                            cell.value = "'" + cell.value

                # Set column widths based on heading length
                for col_idx, cell in enumerate(worksheet[2], start=1):
                    column_letter = get_column_letter(col_idx)
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        worksheet.column_dimensions[column_letter].width = max(cell_length, 10)
                    else:
                        worksheet.column_dimensions[column_letter].width = 10

                # Auto filter from heading row
                last_row = worksheet.max_row
                last_col_letter = get_column_letter(len(sections))
                worksheet.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

                # Freeze panes below the headings
                worksheet.freeze_panes = worksheet['A3']

                # === Vertical borders between section blocks (full height) ===
                last_section = sections[0]
                max_row = worksheet.max_row
                for col_idx in range(2, len(sections) + 1):
                    if sections[col_idx - 1] != last_section:
                        for row_idx in range(1, max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.border = Border(
                                left=thin_side,
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=cell.border.bottom
                            )
                        last_section = sections[col_idx - 1]

        tqdm.write(f"Saved data to '{output_filename}.xlsx'")

    except Exception as e:
        tqdm.write(f"Saving data to '{output_filename}.xlsx' failed. Error: {e}")

# Helper function to merge and center cells in the first row of the worksheet
def merge_and_center_cells(worksheet, sections):
    """
    Merges and centers identical consecutive section values in the first row of the worksheet.
    Args:
        worksheet: The worksheet object where merging is applied.
        sections: A list of section titles corresponding to the columns.
    """
    start_col = 1
    for col_idx in range(1, len(sections) + 1):
        if col_idx == len(sections) or sections[col_idx] != sections[start_col - 1]:
            if col_idx - start_col >= 1:
                worksheet.merge_cells(
                    start_row=1, start_column=start_col,
                    end_row=1, end_column=col_idx
                )
            # Apply formatting to all cells in merged range
            for i in range(start_col, col_idx + 1):
                cell = worksheet.cell(row=1, column=i)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='4F81BD')
            start_col = col_idx + 1
